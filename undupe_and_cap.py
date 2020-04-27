from openpyxl import load_workbook

# NOTE: Rows in the warning message are referring to the rows from the original file, not the newly formatted one
# NOTE: Excel file to be targeted has to be in the same file from which this python program is being run

# Tool to make excel formatting/sorting a little bit easier
#  - Formats names so the first letter is Capped and the rest aren't
#  - removes rows of people who have names that are already in the worksheet
#      - This checks for people with the same name from different companies. If there are people with the same exact
#        name but at different companies, they will be unintentionally deleted.
#  - Outputs the names of the people who were deleted in terminal. This should be used to double check to make sure
#    people weren't unintentionally removed.

# use input() to find the name of the file you want to target.
source_file_name = input("What's the name of the file you want to format?\n(Without \".xlsx\" extension): ")
header_row = input("What row is the header on? (everything above will be ignored): ")
first_name_index = input("What's the column index of First Name? (A = 1 etc.): ")
last_name_index = input("What's the column index of Last Name? (A = 1 etc.): ")
company_index = input("What's the column index of Company Name? (A = 1 etc.): ")

# open that file and call it "source" (don't save over this one)
source_wb = load_workbook(source_file_name + ".xlsx")
source_sheet = source_wb.active

# create a dictionary to store name:company pairs
names_and_companies = {}
counter = int(header_row)
# add to this as I find duplicates
rows_to_be_deleted = []
for row in source_sheet.iter_rows(min_row=int(header_row)):
    # formatting first name
    first_name = str(row[int(first_name_index)].value).capitalize().strip()
    row[int(first_name_index)].value = first_name

    # formatting last name
    last_name = str(row[int(last_name_index)].value).capitalize().strip()
    row[int(last_name_index)].value = last_name

    company = str(row[int(company_index)].value)
    whole_name = first_name + " " + last_name

    # add the name if it's not there already
    # row is kept the same in this case (except for name formatting)
    if whole_name not in names_and_companies:
        names_and_companies[whole_name] = company

    # handle case where name is the same but company isn't
    # row maintained but flag is raised and names are formatted
    elif whole_name in names_and_companies and names_and_companies[whole_name] is not company:
        print("(Original File Row " + str(counter) + ") Possible Duplicate: " + whole_name + " " + company + " where " + whole_name + " " +
              names_and_companies[whole_name] + " already exists! (ADDED ANYWAYS)")

    # print out alert if name and company are the same
    # row marked for termination
    elif whole_name in names_and_companies and names_and_companies[whole_name] is company:
        print("(Original File Row " + str(counter) + ") Duplicate Name and Company: " + whole_name + " " +
              names_and_companies[whole_name] + " already exists!")
        rows_to_be_deleted.append(counter)

    counter += 1

# deleting the rows with duplicates being careful to delete from the bottom up so we don't have accidental deletions
rows_to_be_deleted.sort(reverse=True)
for rownum in rows_to_be_deleted:
    source_sheet.delete_rows(rownum)

# this creates a new fil with the original file's name + an extension
source_wb.save(filename=source_file_name + "-formatted.xlsx")
